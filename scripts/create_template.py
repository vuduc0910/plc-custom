from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pathlib import Path


def create_default_template(output_path: str = "./config/template.xlsx") -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    center = Alignment(horizontal="center")

    headers_a = ["Port", "Value"]
    for col, text in enumerate(headers_a, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for port in range(1, 10):
        ws.cell(row=port + 1, column=1, value=f"p{port}")
        ws.cell(row=port + 1, column=2, value=0.0)

    output_headers = {
        "K": "G1 Output",
        "L": "G2 Output",
        "M": "G3 Output",
    }
    for col_letter, title in output_headers.items():
        cell = ws[f"{col_letter}1"]
        cell.value = title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    ws["K2"] = "=(B2+B3+B4+B5)/4"
    ws["L2"] = "=(B6+B7+B8+B9)/4"
    ws["M2"] = "=B10"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 14

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()


if __name__ == "__main__":
    create_default_template()
