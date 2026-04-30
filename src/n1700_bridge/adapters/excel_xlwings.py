"""Excel data source using openpyxl for reading measurement data.

Note: The real implementation would use xlwings for live Excel reading.
For the POC with fakes, we use openpyxl to avoid requiring Excel to be installed.
"""

from datetime import datetime
from pathlib import Path

from loguru import logger
from openpyxl import load_workbook

from n1700_bridge.core.excel_source import ExcelSourceError
from n1700_bridge.core.models import PortReading

_NUM_PORTS = 9


class OpenpyxlExcelSource:
    """Reads measurement data from Excel file using openpyxl.

    Used as the ExcelDataSource implementation when use_fake=true.
    For real deployment, XlwingsExcelSource would connect to a live Excel window.
    """

    def __init__(
        self,
        path: Path,
        sheet_name: str = "Sheet1",
        header_row: int = 1,
        port_columns: list[str] | None = None,
    ) -> None:
        self._path = path
        self._sheet_name = sheet_name
        self._header_row = header_row
        self._port_columns = port_columns or ["B", "C", "D", "E", "F", "G", "H", "I", "J"]

    def read_latest_row(self) -> list[PortReading]:
        """Return readings for ports 1..9 from the last data row.

        Raises:
            ExcelSourceError: If file doesn't exist, is empty, or read fails.
        """
        logger.info("Reading latest row from {}", self._path)

        if not self._path.exists():
            raise ExcelSourceError(f"Excel file not found: {self._path}")

        try:
            wb = load_workbook(self._path, read_only=True, data_only=True)
        except Exception as e:
            raise ExcelSourceError(f"Cannot open Excel file: {e}") from e

        try:
            ws = wb[self._sheet_name]

            # Find last row with data
            last_row = ws.max_row
            if last_row is None or last_row <= self._header_row:
                raise ExcelSourceError("Excel file has no data rows")

            readings: list[PortReading] = []
            for port_idx, col_letter in enumerate(self._port_columns[:_NUM_PORTS], start=1):
                cell = ws[f"{col_letter}{last_row}"]
                val = cell.value
                if val is None:
                    val = 0.0
                readings.append(PortReading(port=port_idx, value=float(val)))

            logger.debug("Read {} readings from row {}", len(readings), last_row)
            return readings
        finally:
            wb.close()

    def get_last_row_timestamp(self) -> datetime:
        """Return the timestamp from column A of the last data row."""
        if not self._path.exists():
            raise ExcelSourceError(f"Excel file not found: {self._path}")

        try:
            wb = load_workbook(self._path, read_only=True, data_only=True)
        except Exception as e:
            raise ExcelSourceError(f"Cannot open Excel file: {e}") from e

        try:
            ws = wb[self._sheet_name]
            last_row = ws.max_row
            if last_row is None or last_row <= self._header_row:
                raise ExcelSourceError("Excel file has no data rows")

            val = ws[f"A{last_row}"].value
            if isinstance(val, datetime):
                return val
            if isinstance(val, str):
                return datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
            return datetime.now()
        finally:
            wb.close()

    def is_file_available(self) -> bool:
        """Check if the Excel file exists and is readable."""
        return self._path.exists()
