"""Fake N1700 controller that appends random readings to Excel file."""

import random
from pathlib import Path

from loguru import logger
from openpyxl import Workbook, load_workbook

_NUM_PORTS = 9


class FakeN1700Controller:
    """Simulates N1700 app by appending random measurement rows to Excel.

    Each click_data_button() call appends a new row with 9 plausible readings
    drifting from previous values. Uses openpyxl (NOT xlwings) to avoid lock conflict.
    """

    def __init__(self, excel_path: Path) -> None:
        self._excel_path = excel_path
        self._last_values: list[float] = [0.0] * _NUM_PORTS
        self._window_available = True

    def click_data_button(self) -> None:
        """Append a new row of random-ish readings to the Excel file."""
        logger.info("FakeN1700 click_data_button -> appending row to {}", self._excel_path)

        # Generate drifting values
        new_values: list[float] = []
        for prev in self._last_values:
            drift = random.gauss(0, 0.02)
            new_values.append(round(prev + drift, 6))
        self._last_values = new_values

        # Append to Excel
        if self._excel_path.exists():
            wb = load_workbook(self._excel_path)
        else:
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Sheet1"
                ws.append(["Timestamp"] + [f"Port{i}" for i in range(1, _NUM_PORTS + 1)])

        ws = wb.active
        if ws is not None:
            from datetime import datetime

            ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S")] + new_values)
        wb.save(self._excel_path)
        wb.close()

        logger.debug("FakeN1700 appended values: {}", new_values)

    def is_window_available(self) -> bool:
        """Always available in fake mode."""
        return self._window_available
