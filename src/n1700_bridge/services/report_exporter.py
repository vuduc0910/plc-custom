from datetime import datetime
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from n1700_bridge.core.models import Measurement, Verdict

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center")
_OK_FILL = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
_NG_FILL = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
_VERDICT_FONT = Font(bold=True, color="FFFFFF", size=7)
_NUMBER_FORMAT = "0.00"
_DECIMALS = 2

# Port groupings
_W_PORTS = [1, 2, 3, 4]
_G_PORTS = [5, 6, 7, 8]
_B_PORT = 9

# Column layout (1-indexed):
#  1  Timestamp
#  2  Part ID
#  W group (13 cols, 3-15):
#   3  W1   4  WV1   5  W2   6  WV2   7  W3   8  WV3   9  W4  10  WV4
#  11  Wmax 12 Wmin 13 Wmax-min 14 W-Ave 15 W-Result
#  G group (13 cols, 16-28):
#  16  G1  17  GV1  18  G2  19  GV2  20  G3  21  GV3  22  G4  23  GV4
#  24  Gmax 25 Gmin 26 Gmax-min 27 G-Ave 28 G-Result
#  Bottom (2 cols, 29-30):
#  29  Bottom  30  B-Result

_W_START = 3
_G_START = 16
_B_START = 29

# Verdict columns (for colour styling and compact width)
_VERDICT_COLS = {4, 6, 8, 10, 15, 17, 19, 21, 23, 28, 30}
# Numeric value columns
_NUMERIC_COLS = {3, 5, 7, 9, 11, 12, 13, 14, 16, 18, 20, 22, 24, 25, 26, 27, 29}


class ReportExporter:

    def __init__(self, output_dir: Path) -> None:
        self._output_dir = output_dir

    def export(self, measurements: list[Measurement]) -> Path:
        self._output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filepath = self._output_dir / f"{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Report")
        else:
            ws.title = "Report"

        headers = _build_headers()
        ws.append(headers)
        _style_header_row(ws, len(headers))

        for m in measurements:
            ws.append(_build_data_row(m))

        _style_verdict_cells(ws, len(measurements))
        _style_numeric_cells(ws, len(measurements))
        _set_column_widths(ws)

        wb.save(filepath)
        wb.close()

        logger.info("Report exported: {} ({} measurements)", filepath, len(measurements))
        return filepath


def _build_headers() -> list[str]:
    return [
        "Timestamp", "Part ID",
        # W group — ports 1-4
        "W1", "WV1", "W2", "WV2", "W3", "WV3", "W4", "WV4",
        "Wmax", "Wmin", "Wmax-min", "W-Ave", "W-Result",
        # G group — ports 5-8
        "G1", "GV1", "G2", "GV2", "G3", "GV3", "G4", "GV4",
        "Gmax", "Gmin", "Gmax-min", "G-Ave", "G-Result",
        # Bottom — port 9
        "Bottom", "B-Result",
    ]


def _build_data_row(m: Measurement) -> list[object]:
    readings = {r.port: round(r.value, _DECIMALS) for r in m.readings}
    verdicts = {pv.port: pv.verdict.value for pv in m.port_verdicts}

    def val(port: int) -> float:
        return readings.get(port, 0.0)

    def vdict(port: int) -> str:
        return verdicts.get(port, "")

    def group_stats(ports: list[int]) -> tuple[float, float, float, float]:
        values = [val(p) for p in ports]
        vmax = round(max(values), _DECIMALS)
        vmin = round(min(values), _DECIMALS)
        return vmax, vmin, round(vmax - vmin, _DECIMALS), round(sum(values) / len(values), _DECIMALS)

    def group_result(ports: list[int]) -> str:
        if all(verdicts.get(p) == Verdict.OK.value for p in ports):
            return Verdict.OK.value
        return Verdict.NG.value

    w_max, w_min, w_range, w_avg = group_stats(_W_PORTS)
    g_max, g_min, g_range, g_avg = group_stats(_G_PORTS)

    return [
        m.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        m.part_id,
        # W group
        val(1), vdict(1), val(2), vdict(2), val(3), vdict(3), val(4), vdict(4),
        w_max, w_min, w_range, w_avg, group_result(_W_PORTS),
        # G group
        val(5), vdict(5), val(6), vdict(6), val(7), vdict(7), val(8), vdict(8),
        g_max, g_min, g_range, g_avg, group_result(_G_PORTS),
        # Bottom
        val(9), vdict(9),
    ]


def _style_header_row(ws: object, col_count: int) -> None:
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)  # type: ignore[union-attr]
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _style_numeric_cells(ws: object, row_count: int) -> None:
    for row_idx in range(2, row_count + 2):
        for col in _NUMERIC_COLS:
            cell = ws.cell(row=row_idx, column=col)  # type: ignore[union-attr]
            if isinstance(cell.value, (int, float)):
                cell.number_format = _NUMBER_FORMAT


def _style_verdict_cells(ws: object, row_count: int) -> None:
    for row_idx in range(2, row_count + 2):
        for col in _VERDICT_COLS:
            cell = ws.cell(row=row_idx, column=col)  # type: ignore[union-attr]
            if cell.value == Verdict.OK.value:
                cell.fill = _OK_FILL
                cell.font = _VERDICT_FONT
            elif cell.value == Verdict.NG.value:
                cell.fill = _NG_FILL
                cell.font = _VERDICT_FONT
            cell.alignment = _HEADER_ALIGN


_COL_WIDTHS: dict[int, float] = {
    1: 21.0,    # Timestamp
    2: 14.0,    # Part ID
    3: 8.0,     # W1
    4: 4.0,     # WV1
    5: 8.0,     # W2
    6: 4.0,     # WV2
    7: 8.0,     # W3
    8: 4.0,     # WV3
    9: 8.0,     # W4
    10: 4.0,    # WV4
    11: 6.8,    # Wmax
    12: 6.0,    # Wmin
    13: 10.2,   # Wmax-min
    14: 10.2,   # W-Ave
    15: 8.0,    # W-Result
    16: 6.0,    # G1
    17: 4.0,    # GV1
    18: 8.0,    # G2
    19: 4.0,    # GV2
    20: 8.0,    # G3
    21: 4.0,    # GV3
    22: 8.0,    # G4
    23: 4.0,    # GV4
    24: 6.3,    # Gmax
    25: 6.0,    # Gmin
    26: 10.2,   # Gmax-min
    27: 6.2,    # G-Ave
    28: 8.3,    # G-Result
    29: 8.0,    # Bottom
    30: 7.9,    # B-Result
}


def _set_column_widths(ws: object) -> None:
    for col_idx, width in _COL_WIDTHS.items():
        col_letter = ws.cell(row=1, column=col_idx).column_letter  # type: ignore[union-attr]
        ws.column_dimensions[col_letter].width = width  # type: ignore[union-attr]
