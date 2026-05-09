from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from n1700_bridge.core.models import (
    JudgmentGroup,
    Measurement,
    PortReading,
    Verdict,
)
from n1700_bridge.services.report_exporter import ReportExporter


def _make_measurement(
    part_id: str = "TEST001",
    base_value: float = 0.01,
    verdict: Verdict = Verdict.OK,
) -> Measurement:
    readings = [
        PortReading(port=i, value=base_value * i) for i in range(1, 10)
    ]
    judgments = [
        JudgmentGroup(
            group_name="G1", output_cell="K2",
            computed_value=base_value * 2.5, verdict=verdict,
        ),
        JudgmentGroup(
            group_name="G2", output_cell="L2",
            computed_value=base_value * 5.5, verdict=verdict,
        ),
        JudgmentGroup(
            group_name="G3", output_cell="M2",
            computed_value=base_value * 8.5, verdict=verdict,
        ),
    ]
    return Measurement(
        timestamp=datetime(2025, 1, 15, 10, 30, 0),
        part_id=part_id,
        readings=readings,
        judgments=judgments,
    )


class TestReportExporter:

    def test_export_creates_file(self, tmp_path: Path) -> None:
        exporter = ReportExporter(tmp_path)
        result = exporter.export([_make_measurement()])

        assert result.exists()
        assert result.suffix == ".xlsx"
        assert result.parent == tmp_path

    def test_export_file_has_correct_headers(self, tmp_path: Path) -> None:
        exporter = ReportExporter(tmp_path)
        result = exporter.export([_make_measurement()])
        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None

        headers = [ws.cell(row=1, column=i).value for i in range(1, 21)]
        expected = (
            ["Timestamp", "Part ID"]
            + [f"Port {i}" for i in range(1, 10)]
            + ["G1 Cell", "G1 Value", "G1 Verdict"]
            + ["G2 Cell", "G2 Value", "G2 Verdict"]
            + ["G3 Cell", "G3 Value", "G3 Verdict"]
        )
        assert headers == expected
        wb.close()

    def test_export_correct_row_count(self, tmp_path: Path) -> None:
        exporter = ReportExporter(tmp_path)
        measurements = [_make_measurement(part_id=f"PART{i}") for i in range(10)]
        result = exporter.export(measurements)
        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None

        assert ws.max_row == 11
        wb.close()

    def test_export_row_data(self, tmp_path: Path) -> None:
        exporter = ReportExporter(tmp_path)
        measurement = _make_measurement(part_id="ABC123", base_value=0.02)
        result = exporter.export([measurement])

        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None

        assert ws.cell(row=2, column=1).value == "2025-01-15 10:30:00"
        assert ws.cell(row=2, column=2).value == "ABC123"
        assert ws.cell(row=2, column=3).value == 0.02
        assert ws.cell(row=2, column=11).value == 0.18

        assert ws.cell(row=2, column=12).value == "K2"
        assert ws.cell(row=2, column=14).value == "OK"
        wb.close()

    def test_export_ng_verdicts(self, tmp_path: Path) -> None:
        exporter = ReportExporter(tmp_path)
        measurement = _make_measurement(verdict=Verdict.NG)
        result = exporter.export([measurement])

        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None

        assert ws.cell(row=2, column=14).value == "NG"
        assert ws.cell(row=2, column=17).value == "NG"
        assert ws.cell(row=2, column=20).value == "NG"
        wb.close()

    def test_export_empty_list(self, tmp_path: Path) -> None:
        exporter = ReportExporter(tmp_path)
        result = exporter.export([])

        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None

        assert ws.max_row == 1
        wb.close()

    def test_export_creates_output_dir(self, tmp_path: Path) -> None:
        output_dir = tmp_path / "nested" / "reports"
        exporter = ReportExporter(output_dir)
        result = exporter.export([_make_measurement()])

        assert output_dir.exists()
        assert result.exists()
